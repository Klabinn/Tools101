import os
import openpyxl
import random

def list_files_to_excel(directory, excel_file):

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the title of the worksheet
    sheet.title = "File List"

    # first row
    sheet.cell(row=1, column=1, value="File Name")

    # Iterate next row
    row = 2
    for filename in os.listdir(directory):
        if os.path.isfile(os.path.join(directory, filename)):
            sheet.cell(row=row, column=1, value=filename)
            row += 1

    workbook.save(excel_file)
    print(f"File names from {directory} have been written to {excel_file}")

# Variabel
directory = './'
excel_file = f"{random.randint(0, 6969)}_scraper.xlsx"

# Call the function
list_files_to_excel(directory, excel_file)