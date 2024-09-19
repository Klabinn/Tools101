import os
import openpyxl
import random

def list_files_to_excel(root_directory, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "File List"

    sheet.cell(row=1, column=1, value="File Name")
    sheet.cell(row=1, column=2, value="Directory Path")

    excluded_files = ['scraper-revamp.py']

    file_count = {}
    original_filenames = {}
    file_paths = {}

    # Dir walk ke smua dir
    row = 2
    for dirpath, dirnames, filenames in os.walk(root_directory):
        for filename in filenames:
            
            normalized_filename = filename.strip().lower()

            if filename in excluded_files:
                continue

            if normalized_filename in file_count:
                file_count[normalized_filename] += 1
                file_paths[normalized_filename].append(dirpath)
                continue
            else:
                file_count[normalized_filename] = 1
                original_filenames[normalized_filename] = filename
                file_paths[normalized_filename] = [dirpath]

            # Format filepathnya
            file_path = os.path.join(dirpath, filename)
            # Printf
            sheet.cell(row=row, column=1, value=filename)
            sheet.cell(row=row, column=2, value=file_path)
            row += 1

    workbook.save(excel_file)
    print(f"File names from {root_directory} and its subdirectories have been written to {excel_file}")

    # Jumlah Dupe
    total_duplicates = sum(count - 1 for count in file_count.values() if count > 1)
    print(f"\nTotal number of duplicate files: {total_duplicates}")

    # Print dupe path, and filenames
    print("\nDuplicate files and their counts:")
    for normalized_filename, count in file_count.items():
        if count > 1:
            original_name = original_filenames[normalized_filename]
            print(f"{original_name}: {count} times")
            print("  Found in directories:")
            for path in file_paths[normalized_filename]:
                print(f"    - {path}")

# Variables
directory = './'
excel_file = f"{random.randint(0, 6969)}_scraper.xlsx"

# Call the function
list_files_to_excel(directory, excel_file)
